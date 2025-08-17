# -*- coding: utf-8 -*-

from sklearn.model_selection import train_test_split
import pandas as pd
import torch.optim as optim
from ignite.engine import Engine, Events, create_supervised_evaluator
from ignite.metrics import Loss
from ignite.contrib.metrics.regression import R2Score
import torch
import torch.nn as nn
from torch.utils.data import Dataset
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl import Workbook

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Transfer learning
# Loading data
dataset_train_tt = pd.read_excel('TL_data_target_task_train.xlsx', engine='openpyxl')
dataset_train_tt = dataset_train_tt.dropna()
num_columns_train = len(dataset_train_tt.columns)

X_train_tt = dataset_train_tt.iloc[:, 1:num_columns_train-1].values
y_train_tt = dataset_train_tt.iloc[:, -1].values

scaler = StandardScaler()
scaler.fit(X_train_tt)
X_train_tt = scaler.transform(X_train_tt)

y_train_tt = np.array(y_train_tt).reshape(len(y_train_tt),1)

X_train_tt_1, X_val_tt, y_train_tt_1, y_val_tt = train_test_split(
                                                              X_train_tt, 
                                                              y_train_tt, 
                                                              test_size = 0.1, 
                                                              random_state = 42,
                                                              )

train_set_tt = pd.concat([pd.DataFrame(X_train_tt_1), pd.DataFrame(y_train_tt_1)], axis = 1)
val_set_tt = pd.concat([pd.DataFrame(X_val_tt), pd.DataFrame(y_val_tt)], axis = 1)

# Build the neural network model
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, hidden_size3, output_size):
        super(NeuralNet, self).__init__()
        self.input_layer = nn.Linear(input_size, hidden_size1)
        self.hidden_layer1 = nn.Linear(hidden_size1, hidden_size2)
        self.hidden_layer2 = nn.Linear(hidden_size2, hidden_size3)
        self.output_layer = nn.Linear(hidden_size3, output_size)
        self.acti = nn.ReLU()

    def forward(self, x):
        x = self.acti(self.input_layer(x))
        x = self.acti(self.hidden_layer1(x))
        x = self.acti(self.hidden_layer2(x))
        x = self.output_layer(x)
        return x   

input_size = X_train_tt_1.shape[1]
hidden_size1 = 64
hidden_size2 = 32
hidden_size3 = 16
output_size = 1

model = NeuralNet(input_size, hidden_size1, hidden_size2, hidden_size3, output_size)

# Use the pretrained model
trained_model = torch.load('Pretrained_model.ckpt')
print(trained_model)

class MyDataset(Dataset):
    def __init__(self, dataframe, data_col_start, data_col_end, labels_col):
        x = dataframe.iloc[:, data_col_start:data_col_end].values
        y = dataframe.iloc[:, labels_col].values
        
        self.x = torch.tensor(x, dtype=torch.float32)
        y = torch.tensor(y, dtype=torch.float32)
        self.y = y.view(-1,1)

    def __len__(self):
        return len(self.y)
    
    def __getitem__(self, index):  # Corrected __getitem__ method
        return self.x[index], self.y[index]

train_set_1_tt = MyDataset(train_set_tt, 0, 255, 255)
val_set_1_tt =  MyDataset(val_set_tt, 0, 255, 255)
batch_size = 32

# Set up train set and val set
train_loader_tt = torch.utils.data.DataLoader(train_set_1_tt, batch_size = batch_size)
val_loader_tt = torch.utils.data.DataLoader(val_set_1_tt, batch_size = len(val_set_1_tt))

# Freeze/unfreeze the parameters of the pre-trained model
for param in trained_model.input_layer.parameters():
    param.requires_grad = False
for param in trained_model.hidden_layer1.parameters():
    param.requires_grad = True
for param in trained_model.hidden_layer2.parameters():
    param.requires_grad = False
for param in trained_model.output_layer.parameters():
    param.requires_grad = False
    
# Define new optimizer
criterion = nn.MSELoss()
metrics_tt = {"loss": Loss(criterion), "r_2": R2Score()}
optimizer_tt = optim.Adam(
                          filter(lambda p: p.requires_grad, trained_model.parameters()), 
                          lr = 0.0001,
                          )

# Training loop
def train_step_1(engine, batch):
    x, y = batch
    trained_model.train()
    optimizer_tt.zero_grad()
    y_pred = trained_model(x)
    loss = criterion(y_pred, y)
    loss.backward()
    optimizer_tt.step()
    return loss.item()

# Build train engine
transfer_trainer = Engine(train_step_1)

train_evaluator = create_supervised_evaluator(trained_model, metrics = metrics_tt)
val_evaluator = create_supervised_evaluator(trained_model, metrics = metrics_tt)

train_loss_tt = []
train_r_2_tt = []
val_loss_tt = []
val_r_2_tt = []

@transfer_trainer.on(Events.EPOCH_COMPLETED(every=10))
def store_metrics(engine):
    train_evaluator.run(train_loader_tt)
    val_evaluator.run(val_loader_tt)
    out = train_evaluator.state.metrics
    out_2 = val_evaluator.state.metrics
    train_loss_tt.append(out["loss"]) 
    train_r_2_tt.append(out["r_2"])
    val_loss_tt.append(out_2["loss"])
    val_r_2_tt.append(out_2["r_2"])
 
train_loader_tt = torch.utils.data.DataLoader(train_set_1_tt, batch_size = batch_size, shuffle = True)
transfer_trainer.run(train_loader_tt, max_epochs = 1000)

# Attach metrics to the trainer and validation evaluator
for name, metric in metrics_tt.items():
    metric.attach(transfer_trainer, name)

dic = {
       "train_loss": train_loss_tt,
       "train_r_2": train_r_2_tt,
       "val_loss": val_loss_tt,
       "val_r_2": val_r_2_tt,
      }
#filename = f"data_tl_epochs-{1}_bs-{10000}"
#case_metrics = pd.DataFrame.from_dict(dic)
#case_metrics.to_json(f"{filename}.json")


# Prediction
trained_model.eval()

# Load dataset
dataset_test_tt = pd.read_excel('TL_data_target_task_test.xlsx', engine='openpyxl')
num_columns_test = len(dataset_test_tt.columns)
MOF_name = dataset_test_tt.iloc[:, 0].values
X_tt_test = dataset_test_tt.iloc[:, 1:num_columns_test-1].values
scaler = StandardScaler()
scaler.fit(X_tt_test)
X_tt_test = scaler.transform(X_tt_test)

data_X_tt = torch.tensor(X_tt_test, dtype=torch.float32)

with torch.no_grad():
     y_pred_tt = trained_model(data_X_tt)
y_pred_tt = y_pred_tt.numpy()
y_pred_tt_1 = y_pred_tt.tolist()

# Inputing adsorption amount to excel
# excel
path_excel = 'TL_data_target_task_test.xlsx'
df = pd.read_excel(path_excel, engine='openpyxl')
row_count = len(df)
column_count = len(df.columns)

book = load_workbook(path_excel)
ws = book['Sheet']

MOFname_data = np.hstack((MOF_name.reshape(-1, 1), y_pred_tt_1))
number = len(MOFname_data)
for i in range(number):
    MOF_single = MOFname_data[i]
    MOF_single_copy = np.copy(MOF_single)
    MOF_single_copy.tolist()
    MOFname = MOF_single_copy[0]
    MOF_TSN = MOF_single_copy[1]
    for j in range(row_count):
        name = df.iloc[j, 0]
        if name == MOFname:
            ws.cell(row = j + 2, column = column_count).value = MOF_TSN
            break

book.save(path_excel)

# Choose top 20% MOFs to gcmc and verify the DNN model
MOF_index = np.argsort(y_pred_tt_1, axis = 0)[::-1]
choose_number = round(0.2 * len(MOF_index))
MOF_choosed = []
for j in range(choose_number):
    single_choosed = str(MOF_name[MOF_index[j]])
    MOF_choosed.append(single_choosed)
    
# Extracting the name of these MOFs which need to be calculated EQeq  
#wb = Workbook()
#ws = wb.worksheets[0]
#ws.cell(row = 1, column = 1).value = "Name" # row label
written_content = [] 
header = ['Name']
for k in range(len(MOF_choosed)):
    written_content.append(MOF_choosed[k].split("'")[1].split('.')[0])
output_file = 'MOF_verify_model.csv'
df = pd.DataFrame(written_content, columns = header)
df.to_csv(output_file, index = False, encoding = 'utf-8')
print('The script Transferlearning_finetune.py has completed.')