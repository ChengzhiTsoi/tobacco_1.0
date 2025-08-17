# -*- coding: utf-8 -*-

from sklearn.model_selection import train_test_split
import pandas as pd
from sklearn import preprocessing
import torch.optim as optim
from ignite.engine import Engine, Events, create_supervised_evaluator
from ignite.metrics import Loss
from ignite.contrib.metrics.regression import R2Score
import torch
import torch.nn as nn
from torch.utils.data import Dataset
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl import Workbook

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

# Load dataset
dataset_test_tt = pd.read_excel('TL_data_target_task_test.xlsx', engine='openpyxl')
num_columns_test = len(dataset_test_tt.columns)
MOF_name = dataset_test_tt.iloc[:, 0].values
X_tt_test = dataset_test_tt.iloc[:, 1:num_columns_test-1].values
scaler = StandardScaler()
scaler.fit(X_tt_test)
X_tt_test = scaler.transform(X_tt_test)
data_X_tt = torch.tensor(X_tt_test, dtype=torch.float32)

# Load model
# Build the neural network model
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, hidden_size3, output_size):
        super(NeuralNet, self).__init__()
        self.input_layer = nn.Linear(input_size, hidden_size1)
        self.hidden_layer1 = nn.Linear(hidden_size1, hidden_size2)
        self.hidden_layer2 = nn.Linear(hidden_size2, hidden_size3)
        self.output_layer = nn.Linear(hidden_size3, output_size)
        self.acti = nn.ReLU()

    def forward(self, x):
        x = self.acti(self.input_layer(x))
        x = self.acti(self.hidden_layer1(x))
        x = self.acti(self.hidden_layer2(x))
        x = self.output_layer(x)
        return x   

input_size = X_tt_test.shape[1]
hidden_size1 = 64
hidden_size2 = 32
hidden_size3 = 16
output_size = 1

model = NeuralNet(input_size, hidden_size1, hidden_size2, hidden_size3, output_size)

trained_model = torch.load('Pretrained_model.ckpt')
print(trained_model)
trained_model.eval()

with torch.no_grad():
     y_pred_tt = trained_model(data_X_tt)
y_pred_tt = y_pred_tt.numpy()
y_pred_tt_1 = y_pred_tt.tolist()

# Inputing adsorption amount to excel
# excel
path_excel = 'TL_data_target_task_test.xlsx'
df = pd.read_excel(path_excel, engine='openpyxl')
row_count = len(df)
column_count = len(df.columns)

book = load_workbook(path_excel)
ws = book['Sheet']

MOFname_data = np.hstack((MOF_name.reshape(-1, 1), y_pred_tt_1))
number = len(MOFname_data)
for i in range(number):
    MOF_single = MOFname_data[i]
    MOF_single_copy = np.copy(MOF_single)
    MOF_single_copy.tolist()
    MOFname = MOF_single_copy[0]
    MOF_TSN = MOF_single_copy[1]
    for j in range(row_count):
        name = df.iloc[j, 0]
        if name == MOFname:
            ws.cell(row = j + 2, column = column_count).value = MOF_TSN
            break

book.save(path_excel)
print('The script Transferlearning_originaldnn.py has completed.')