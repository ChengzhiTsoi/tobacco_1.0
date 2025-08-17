# -*- coding: utf-8 -*-

import numpy as np
import pandas as pd
from openpyxl import load_workbook

# txt
with open('N.txt') as e:
    e1 = e.readlines()
number = len(e1)
i = 0

# excel
path_excel = 'TL_data_target_task_train.xlsx'
df = pd.read_excel(path_excel, header=None, engine='openpyxl')
row_count = len(df)
column_count = df.iloc[0].count()
book = load_workbook(path_excel)
ws = book['Sheet']

while i in range(number - 5):
    if 'best_mol_' in e1[i]:
        MOFname = e1[i].split('\n')[0]
        N_CH4 = e1[i + 1]
        N_C2 = e1[i + 2]
        N_C3 = e1[i + 3]
        N_CO2 = e1[i + 4]
        N_H2S = e1[i + 5]
        
        N1 = float(N_CH4.split()[5])
        N2 = float(N_C2.split()[5])
        N3 = float(N_C3.split()[5])
        N4 = float(N_CO2.split()[5])
        N5 = float(N_H2S.split()[5])
        
        MOFname_full = MOFname + '.cif'
        N_total = N4 + N5
        S_total = ((N4 + N5)/0.15) / ((N1 + N2 + N3)/0.85)
        TSN = (N_total) * np.log10(S_total)
        
        for j in range(row_count):
            name = df.iloc[j, 0]
            if name == MOFname_full:
                ws.cell(row = j + 1, column = column_count + 1).value = TSN
                break
            else:
                pass
        
        book.save(path_excel)
        
        i = i + 6
    else:
        i = i + 1 
        pass