# -*- coding: utf-8 -*-

import pandas as pd
import torch
import torch.nn as nn
import numpy as np
from sklearn.preprocessing import StandardScaler
from openpyxl import load_workbook
from openpyxl import Workbook
from EncodeMOFs import EM
import os
import shutil

device = torch.device("cuda" if torch.cuda.is_available() else "cpu")

def delete_unwanted_files_and_folders(directory):
    for p0 in os.listdir(directory):
        full_path = os.path.join(directory, p0)
        if not ('.py' in p0 or 'lammps_input.in' in p0):
            if os.path.isdir(full_path):
                try:
                    shutil.rmtree(full_path)
                except Exception as e:
                    pass
            else:
                try:
                    os.remove(full_path)
                except Exception as e:
                    pass
            
path_gcmc_mofs = 'gcmc_selected_mofs'
delete_unwanted_files_and_folders(path_gcmc_mofs)


output_folder = 'LINUX'
for s in os.listdir(output_folder):
    full_path = os.path.join(output_folder, s)
    if s == 'simulation.input':
        continue
    if s == 'EQeq':
        continue
    if s == 'gcmc_required':
        continue
    if s == 'prepared_mofs':
        if os.path.isdir(full_path):
            for f in os.listdir(full_path):
                fp = os.path.join(full_path, f)
                try:
                    if os.path.isfile(fp) or os.path.islink(fp):
                        os.remove(fp)
                    elif os.path.isdir(fp):
                        shutil.rmtree(fp)
                except Exception as e:
                    print(f"Warning: failed to delete {fp}: {e}")
        continue
    try:
        if os.path.isdir(full_path):
            shutil.rmtree(full_path)
        else:
            os.remove(full_path)
    except Exception as e:
        print(f"Warning: failed to delete {full_path}: {e}")


output_folder1 = 'TL'
for t in os.listdir(output_folder1):
    file_path = os.path.join(output_folder1, t)
    if 'TL_data_target_task_train.xlsx' in t:
        os.remove(file_path)
    elif 'TL_data_target_task_test.xlsx' in t:
        os.remove(file_path)
    elif 'MOF_verify_model.csv' in t:
        os.remove(file_path)
    elif 'Output.png' in t:
        os.remove(file_path)

mof_verify_folder = os.path.join(output_folder1, 'MOF_verify')
if os.path.exists(mof_verify_folder) and os.path.isdir(mof_verify_folder):
    shutil.rmtree(mof_verify_folder)
                        

excel_path = 'TL/final_data.xlsx'
wb = load_workbook(excel_path)
keep_sheet_name = 'Cycle 1'
sheets = wb.sheetnames
for sheet in sheets:
    if sheet != keep_sheet_name:
        wb.remove(wb[sheet])
wb.save(excel_path)


path_test_data = 'gcmc_rest_mofs'
for q0 in os.listdir(path_test_data):
    os.remove(path_test_data + '/' + q0)


# 7. Encoding new MOFs and prepare the TL data
path_new_mof_1 = 'new_designed_mofs'
path_mol = 'best_edges_mol'
train_test=0.1

X_test, sub_name_node, sub_name_topo, linker_fp2, MOF_name, bbb = EM(path_new_mof_1, path_mol)

wb = Workbook() # Inputing data into sheet
ws = wb.worksheets[0]

ws.cell(row = 1, column = 1).value = "Name" # row label
number_FP = len(linker_fp2[1])
for i in range(number_FP):
    singleFP_index = "FP" + str(i)
    ws.cell(row = 1, column = i + 2).value = singleFP_index
for i in range(len(sub_name_node)):
    single_node_index = sub_name_node[i]
    ws.cell(row = 1, column = int(i) + int(number_FP) + int(2)).value = single_node_index
for i in range(len(sub_name_topo)):
    single_topo_index = sub_name_topo[i]
    ws.cell(row = 1, column = int(len(sub_name_node)) + int(i) + int(number_FP) + int(2)).value = single_topo_index
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(2) + int(number_FP)).value = "LengthA"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(3) + int(number_FP)).value = "LengthB"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(4) + int(number_FP)).value = "LengthC"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(5) + int(number_FP)).value = "TSN"

for i in range(len(MOF_name)): # Inputing content
    ws.cell(row = i + 2, column = 1).value = MOF_name[i]
    for j in range(len(bbb[i])):
        ws.cell(row = i + 2, column = 2 + j).value = bbb[i][j]
        
output_file = 'new_designed_mofs/All_designed_mofs.xlsx'
wb.save(output_file)


# 8. Original DNN model was used to predict the performance of all new MOFs
# Use the original model
class NeuralNet(nn.Module):
    def __init__(self, input_size, hidden_size1, hidden_size2, hidden_size3, output_size):
        super(NeuralNet, self).__init__()
        self.input_layer = nn.Linear(input_size, hidden_size1)
        self.hidden_layer1 = nn.Linear(hidden_size1, hidden_size2)
        self.hidden_layer2 = nn.Linear(hidden_size2, hidden_size3)
        self.output_layer = nn.Linear(hidden_size3, output_size)
        self.acti = nn.ReLU()

    def forward(self, x):
        x = self.acti(self.input_layer(x))
        x = self.acti(self.hidden_layer1(x))
        x = self.acti(self.hidden_layer2(x))
        x = self.output_layer(x)
        return x   

trained_model = torch.load('TL/Pretrained_model.ckpt')
trained_model.eval()

# Loading data
data = pd.read_excel('new_designed_mofs/All_designed_mofs.xlsx', engine='openpyxl')
num_columns = len(data.columns)
MOF_name = data.iloc[:, 0].values
X = data.iloc[:, 1:num_columns-1].values
X = X.astype(np.float64)
scaler = StandardScaler()
scaler.fit(X)
X = scaler.transform(X)

X_1 = torch.tensor(X, dtype=torch.float32)
with torch.no_grad():
     y_pred = trained_model(X_1)
     
y_pred = y_pred.numpy()
y_pred_1 = y_pred.tolist()

# Inputing adsorption amount to excel
# excel
path_excel = 'new_designed_mofs/All_designed_mofs.xlsx'
df = pd.read_excel(path_excel, engine='openpyxl')
row_count = len(df)
column_count = len(df.columns)

book = load_workbook(path_excel)
ws = book.worksheets[0]

MOFname_data = np.hstack((MOF_name.reshape(-1, 1), y_pred_1))
number = len(MOFname_data)
for i in range(number):
    MOF_single = MOFname_data[i]
    MOF_single_copy = np.copy(MOF_single)
    MOF_single_copy.tolist()
    MOFname = MOF_single_copy[0]
    MOF_TSN = MOF_single_copy[1]
    for j in range(row_count):
        name = df.iloc[j, 0]
        if name == MOFname:
            ws.cell(row = j + 2, column = column_count).value = MOF_TSN
            break
        else:
            pass
book.save(path_excel)


# 9. Choosing the MOFs with top 20% TSN to be the training data
new_mof_name_0 = os.listdir(path_new_mof_1)
new_mof_name = []
for j0 in new_mof_name_0:
    if '.cif' not in j0:
        pass
    else:
        new_mof_name.append(j0)
        
MOF_index = np.argsort(y_pred, axis = 0)[::-1]
choose_number = round(train_test * len(MOF_index))
MOF_choosed = []
for j in range(choose_number):
    single_choosed = str(MOF_name[MOF_index[j]])
    single_choosed = single_choosed.split("'")[1]
    MOF_choosed.append(single_choosed) # The chosen new MOFs
MOF_not_selected = [mof_name for mof_name in new_mof_name if mof_name not in MOF_choosed] # The rest of new MOFs

# Moving the chosen MOFs to a new file
for p in MOF_choosed:
    shutil.copy(path_new_mof_1 + '/' + p, path_gcmc_mofs + '/' + p)
    
# Extracting the name of these MOFs which need to be calculated EQeq
written_content = os.listdir(path_gcmc_mofs)
written_content_1 = []
header = ['Name']
for r in written_content:
    if '.cif' not in r:
        pass
    else:
        r1 = r.split('.')[0]
        written_content_1.append(r1)

output_file = 'LINUX/MOFs_gcmc.csv'  
output_file1 = 'gcmc_selected_mofs/MOFs_gcmc.csv'
df1 = pd.DataFrame(written_content_1, columns = header)
df1.to_csv(output_file, index = False, encoding = 'utf-8')
df2 = pd.DataFrame(written_content_1, columns = header)
df2.to_csv(output_file1, index = False, encoding = 'utf-8')

# Train set of target task             
X_test, sub_name_node, sub_name_topo, linker_fp2, MOF_name, bbb = EM(path_gcmc_mofs, path_mol)

wb = Workbook() # Inputing data into sheet
ws = wb.worksheets[0]

ws.cell(row = 1, column = 1).value = "Name" # Row label
number_FP = len(linker_fp2[1])
for i in range(number_FP):
    singleFP_index = "FP" + str(i)
    ws.cell(row = 1, column = i + 2).value = singleFP_index
for i in range(len(sub_name_node)):
    single_node_index = sub_name_node[i]
    ws.cell(row = 1, column = int(i) + int(number_FP) + int(2)).value = single_node_index
for i in range(len(sub_name_topo)):
    single_topo_index = sub_name_topo[i]
    ws.cell(row = 1, column = int(len(sub_name_node)) + int(i) + int(number_FP) + int(2)).value = single_topo_index
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(2) + int(number_FP)).value = "LengthA"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(3) + int(number_FP)).value = "LengthB"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(4) + int(number_FP)).value = "LengthC"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(5) + int(number_FP)).value = "TSN"

for i in range(len(MOF_name)): # Inputing content
    ws.cell(row = i + 2, column = 1).value = MOF_name[i]
    for j in range(len(bbb[i])):
        ws.cell(row = i + 2, column = 2 + j).value = bbb[i][j]
        
output_file = 'TL/TL_data_target_task_train.xlsx'
wb.save(output_file)

# Inputing TSN
df_output_file = pd.read_excel(output_file, engine='openpyxl')
row_count_df_output_file = len(df_output_file)

book1 = load_workbook(output_file)
ws = book1.worksheets[0]
number = len(MOFname_data)
for i in range(number):
    MOF_single = MOFname_data[i]
    MOF_single_copy = np.copy(MOF_single)
    MOF_single_copy.tolist()
    MOFname = MOF_single_copy[0]
    MOF_TSN = MOF_single_copy[1]
    for j in range(row_count_df_output_file):
        name = df_output_file.iloc[j, 0]
        if name == MOFname:
            ws.cell(row = j + 2, column = column_count).value = MOF_TSN
            break
        else:
            pass
book1.save(output_file)


# Moving the not chosen MOFs to a new file
for q in MOF_not_selected:
    shutil.copy(path_new_mof_1 + '/' + q, path_test_data + '/' + q)
    
X_test, sub_name_node, sub_name_topo, linker_fp2, MOF_name, bbb = EM(path_test_data, path_mol)

wb = Workbook() # Inputing data into sheet
ws = wb.worksheets[0]

ws.cell(row = 1, column = 1).value = "Name" # row label
number_FP = len(linker_fp2[1])
for i in range(number_FP):
    singleFP_index = "FP" + str(i)
    ws.cell(row = 1, column = i + 2).value = singleFP_index
for i in range(len(sub_name_node)):
    single_node_index = sub_name_node[i]
    ws.cell(row = 1, column = int(i) + int(number_FP) + int(2)).value = single_node_index
for i in range(len(sub_name_topo)):
    single_topo_index = sub_name_topo[i]
    ws.cell(row = 1, column = int(len(sub_name_node)) + int(i) + int(number_FP) + int(2)).value = single_topo_index
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(2) + int(number_FP)).value = "LengthA"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(3) + int(number_FP)).value = "LengthB"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(4) + int(number_FP)).value = "LengthC"
ws.cell(row = 1, column = int(len(sub_name_topo)) + int(len(sub_name_node)) + int(5) + int(number_FP)).value = "TSN"

for i in range(len(MOF_name)): # Inputing content
    ws.cell(row = i + 2, column = 1).value = MOF_name[i]
    for j in range(len(bbb[i])):
        ws.cell(row = i + 2, column = 2 + j).value = bbb[i][j]
        
output_file = 'TL/TL_data_target_task_test.xlsx'
wb.save(output_file)